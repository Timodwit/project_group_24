"""
Magnetic Tweezers — Section 3: Data Analysis & Interpretation
VU Amsterdam Biophysics Practicum 2026
Authors: fill in your names

Dependencies:  numpy, scipy, pandas, matplotlib
Install:       pip install numpy scipy pandas matplotlib
"""


import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.optimize import brentq, curve_fit
from pathlib import Path

plt.rcParams.update({'font.size': 10, 'figure.dpi': 120})

PRACTICAL   = Path(__file__).parent.parent / "plots" / "practical"
THEORETICAL = Path(__file__).parent.parent / "plots" / "theoretical"
PRACTICAL.mkdir(parents=True, exist_ok=True)
THEORETICAL.mkdir(parents=True, exist_ok=True)

# ------------------------------------------------------------------------------
# 0.  PHYSICAL CONSTANTS
# ------------------------------------------------------------------------------
kB       = 1.38e-23          # J/K
T        = 293.15            # K  (20 °C)
kBT_J    = kB * T            # ~4.1e-21 J
kBT_pNnm = kBT_J * 1e21      # ~4.1 pN·nm  (useful check: 1 kBT ~ 4.1 pN·nm)

eta      = 1e-3              # Pa·s   (water viscosity)
r_m      = 0.5e-6            # m      (bead radius 0.5 um)
gamma    = 6 * np.pi * eta * r_m   # Stokes drag  [N·s/m]

L_p      = 45e-9             # m      (persistence length DNA)
N_bp     = 20_600            # base pairs
L_c_m    = N_bp * 0.34e-9   # m      (contour length)

f_ac     = 10.0              # Hz     (camera acquisition frequency for Part 2)

# ------------------------------------------------------------------------------
# 1.  PHYSICS HELPERS
# ------------------------------------------------------------------------------

def wlc_force(L_ext_m: float) -> float:
    """WLC  Eq. 12:  F(L_ext) [pN].  Marko-Siggia formula."""
    x = np.clip(L_ext_m / L_c_m, 1e-9, 0.99999)
    F_N = (kBT_J / L_p) * (1.0 / (4.0*(1.0 - x)**2) - 0.25 + x)
    return F_N * 1e12  # pN


def wlc_L_ext(F_pN: float) -> float:
    """Invert WLC:  F [pN]  ->  L_ext [um]."""
    F_N = F_pN * 1e-12
    def residual(L):
        return wlc_force(L) * 1e-12 - F_N
    L_m = brentq(residual, 1e-12, 0.99999 * L_c_m)
    return L_m * 1e6   # um


def t_cx(F_pN: float, L_ext_um: float) -> float:
    """Characteristic time  t_{c,x} = γ L_ext / F  [s]   Eq. 7."""
    return gamma * (L_ext_um * 1e-6) / (F_pN * 1e-12)


def var_x_theory(F_pN: float, L_ext_um: float) -> float:
    """Theoretical  <x^2>  [nm²]  from Eq. 8  (axis _|_ B)."""
    return kBT_J * (L_ext_um * 1e-6) / (F_pN * 1e-12) * 1e18


def var_y_theory(F_pN: float, L_ext_um: float) -> float:
    """Theoretical  <y^2>  [nm²]  from Eq. 9  (axis // B, includes bead radius)."""
    return kBT_J * (L_ext_um * 1e-6 + r_m) / (F_pN * 1e-12) * 1e18


def force_eq8(var_nm2: float, L_ext_um: float) -> float:
    """Force from Eq. 8   (use on axis _|_ B):   F = kBT·L_ext / <x^2>   [pN]."""
    return kBT_J * (L_ext_um * 1e-6) / (var_nm2 * 1e-18) * 1e12


def force_eq9(var_nm2: float, L_ext_um: float) -> float:
    """Force from Eq. 9   (use on axis // B):   F = kBT·(L_ext+r) / <y^2>   [pN]."""
    return kBT_J * (L_ext_um * 1e-6 + r_m) / (var_nm2 * 1e-18) * 1e12


def shutter_overestimation(tcx_s: float, tau_sh_s: float) -> float:
    """
    F_meas / F_true = π / (2·arctan(π·t_{c,x}/tau_sh))
    Returns the overestimation factor (≥ 1).  Derived from Eq. 11.
    """
    u = np.pi * tcx_s / tau_sh_s
    return np.pi / (2.0 * np.arctan(u))


def F_z(z_mm):
    """Magnet force-distance calibration curve  [pN]  (Preparatory Q4)."""
    return 22.3811 * np.exp(-1.4578 * z_mm) + 52.2987 * np.exp(-0.6912 * z_mm)


def z_from_F(F_pN: float) -> float:
    """Invert F(z):  force [pN]  ->  magnet position z [mm]."""
    return brentq(lambda z: F_z(z) - F_pN, 0.01, 20.0)


# ------------------------------------------------------------------------------
# 2.  DATA I/O & UTILITIES
# ------------------------------------------------------------------------------

def load_data(path_or_df):
    """
    Read a tab-delimited MT data file, or pass through a DataFrame directly.

    Expected column order:
        frame | time_ms | x1 y1 z1 | x2 y2 z2 | … (positions in um)
        Last two bead triplets are reference beads N1, N2.

    Returns
    -------
    df      : DataFrame with named columns
    n_beads : total number of beads (tethered + 2 reference)
    """
    if isinstance(path_or_df, pd.DataFrame):
        df = path_or_df.copy()
        n_beads = (df.shape[1] - 2) // 3
        return df, n_beads
    raw = pd.read_csv(path_or_df, sep='\t', header=None, comment='#')
    n_beads = (raw.shape[1] - 2) // 3
    # Drop any trailing extra columns beyond the expected triplets
    n_cols = 2 + n_beads * 3
    raw = raw.iloc[:, :n_cols]
    cols = ['frame', 'time_ms']
    for i in range(1, n_beads + 1):
        cols += [f'x{i}', f'y{i}', f'z{i}']
    raw.columns = cols
    return raw, n_beads


def pop_variance(arr: np.ndarray) -> float:
    """Population variance  sigma^2 = (1/N) Σ(xᵢ − x̄)²   (Eq. 13)."""
    return float(np.sum((arr - arr.mean())**2) / len(arr))


def sigma_clip(arr: np.ndarray, n_sigma: float = 5.0):
    """
    Remove spikes beyond n_sigma·std.
    Returns (clipped_array, boolean_mask).
    """
    mu  = arr.mean()
    sig = arr.std()
    mask = np.abs(arr - mu) < n_sigma * sig
    return arr[mask], mask


# ------------------------------------------------------------------------------
# 3.  PART 1 — TRACKING RESOLUTION  (Measurement §1)
# ------------------------------------------------------------------------------

def part1_tracking_resolution(file_200gl: str, file_20gl: str):
    """
    Evaluate tracking noise from stuck reference beads.

    Steps
    -----
    1. Variance (x,y,z) of N1 and N2 separately — both illumination levels.
    2. Drift-corrected variance: N2 − N1  (single-bead noise = result / 2).
    3. Compare single-bead noise to theoretical <x^2> at 6 pN.
    4. Plot position traces.
    """
    L6    = wlc_L_ext(6.0)               # L_ext at 6 pN  [um]
    v6_th = var_x_theory(6.0, L6)        # theoretical <x^2> at 6 pN  [nm²]

    print(f"\n{'='*60}")
    print(f"  PART 1 — TRACKING RESOLUTION")
    print(f"  Reference:  <x^2>_theory at 6 pN = {v6_th:.1f} nm²  "
          f"(L_ext = {L6:.2f} um)")
    print(f"{'='*60}")

    results = {}

    for label, fpath in [('200 GL', file_200gl), ('20 GL', file_20gl)]:
        df, n_beads = load_data(fpath)
        n1 = n_beads - 1   # 1-based index of N1
        n2 = n_beads       # 1-based index of N2

        print(f"\n  -- {label} --")
        print(f"  {'Bead':<6} {'axis':<5} {'sigma^2 [nm²]':>12}  {'σ [nm]':>10}")

        per_bead = {}
        for name, idx in [('N1', n1), ('N2', n2)]:
            variances = {}
            for ax in ('x', 'y', 'z'):
                arr = df[f'{ax}{idx}'].values * 1e3  # um -> nm
                v   = pop_variance(arr)
                variances[ax] = v
                print(f"  {name:<6} {ax:<5} {v:>12.2f}  {np.sqrt(v):>10.2f}")
            per_bead[name] = variances

        # Drift-corrected  N2 − N1
        print(f"\n  Drift-corrected (N2 − N1):")
        print(f"  {'axis':<5} {'sigma^2_corr':>12}  {'sigma^2_single':>12}  "
              f"{'ratio to 6pN theory':>22}")
        drift_corr = {}
        for ax in ('x', 'y', 'z'):
            n1_arr = df[f'{ax}{n1}'].values * 1e3
            n2_arr = df[f'{ax}{n2}'].values * 1e3
            diff   = n2_arr - n1_arr
            v_corr   = pop_variance(diff)
            v_single = v_corr / 2.0   # both beads contribute equally
            ratio    = v_single / v6_th
            drift_corr[ax] = v_single
            print(f"  {ax:<5} {v_corr:>12.2f}  {v_single:>12.2f}  "
                  f"  {ratio:>19.4f}×")

        results[label] = {'per_bead': per_bead, 'drift_corr': drift_corr,
                          'v6_theory': v6_th}

        # -- Plots ----------------------------------------------------------
        t      = df['time_ms'].values / 1e3   # s
        n1x_nm = df[f'x{n1}'].values * 1e3
        n2x_nm = df[f'x{n2}'].values * 1e3
        n1z_nm = df[f'z{n1}'].values * 1e3
        n2z_nm = df[f'z{n2}'].values * 1e3
        diff_x = n2x_nm - n1x_nm

        fig, axes = plt.subplots(3, 1, figsize=(12, 7), sharex=True)

        axes[0].plot(t, n1x_nm, lw=0.6, label='N1', alpha=0.8)
        axes[0].plot(t, n2x_nm, lw=0.6, label='N2', alpha=0.8)
        axes[0].set_ylabel('x  [nm]')
        axes[0].legend(fontsize=8)
        axes[0].set_title(f'Raw x-positions — {label}')

        axes[1].plot(t, diff_x, lw=0.6, color='seagreen', label='N2 − N1')
        axes[1].axhline(0, color='k', lw=0.5, ls='--')
        axes[1].set_ylabel('Δx  [nm]')
        axes[1].legend(fontsize=8)
        axes[1].set_title('Drift-corrected x')

        axes[2].plot(t, n1z_nm, lw=0.6, label='N1', alpha=0.8)
        axes[2].plot(t, n2z_nm, lw=0.6, label='N2', alpha=0.8)
        axes[2].set_ylabel('z  [nm]')
        axes[2].set_xlabel('Time  [s]')
        axes[2].legend(fontsize=8)
        axes[2].set_title('z-positions (drift visible here)')

        plt.suptitle(f'Reference bead traces — {label}', fontsize=11, y=1.01)
        plt.tight_layout()
        out = str(PRACTICAL / f"tracking_{label.replace(' ', '_')}.svg")
        plt.savefig(out, bbox_inches='tight')
        plt.show()
        plt.close()
        print(f"  -> saved {out}")

    return results


# ------------------------------------------------------------------------------
# 4.  PART 2 — FORCE CALIBRATION CORE
# ------------------------------------------------------------------------------

SET_FORCES = [0.1, 0.2, 0.3, 0.4, 0.8, 1.0, 3.0, 5.0, 6.0]   # pN


def _analyze_bead(df: pd.DataFrame,
                  bead_idx: int,
                  ref_idx: int,
                  n_zero: int,
                  t_fail_ms: float = None) -> dict:
    """
    Extract L_ext, variances, and all force estimates for one tethered bead.

    Parameters
    ----------
    df         : DataFrame for one force-step file
    bead_idx   : 1-based index of the tethered bead
    ref_idx    : 1-based index of the reference bead (drift correction)
    n_zero     : number of frames at zero force at the beginning
    t_fail_ms  : discard data after this experiment time [ms] (None = keep all)

    Returns dict with L_ext, variances, four force estimates
    """
    # Drift correction
    for ax in ('x', 'y', 'z'):
        df[f'd{ax}'] = (df[f'{ax}{bead_idx}'] - df[f'{ax}{ref_idx}']) * 1e3  # nm

    # L_ext from z: |median(z_force) − median(z_zero)|
    z_all   = df['dz'].values
    z_zero  = z_all[:n_zero]
    z_force = df['dz'].values[n_zero:]

    # Apply DNA failure cutoff to force region
    if t_fail_ms is not None:
        t_force = df['time_ms'].values[n_zero:]
        valid   = t_force <= t_fail_ms
        z_force = z_force[valid]

    L_ext_nm = abs(np.median(z_force) - np.median(z_zero))
    L_ext_um = L_ext_nm / 1e3

    # Lateral variances in force region (sigma-clip spikes first)
    x_raw = df['dx'].values[n_zero:]
    y_raw = df['dy'].values[n_zero:]

    if t_fail_ms is not None:
        t_force = df['time_ms'].values[n_zero:]
        valid   = t_force <= t_fail_ms
        x_raw   = x_raw[valid]
        y_raw   = y_raw[valid]
    x_c, _ = sigma_clip(x_raw)
    y_c, _ = sigma_clip(y_raw)
    var_x = pop_variance(x_c)   # nm²
    var_y = pop_variance(y_c)   # nm²

    # Force estimates — apply Eq. 8 to both axes
    Fx8 = force_eq8(var_x, L_ext_um)
    Fy8 = force_eq8(var_y, L_ext_um)
    # Eq. 9 accounts for bead radius (correct for the axis // B)
    Fx9 = force_eq9(var_x, L_ext_um)
    Fy9 = force_eq9(var_y, L_ext_um)

    return dict(L_ext_um=L_ext_um,
                var_x=var_x, var_y=var_y,
                Fx_eq8=Fx8, Fy_eq8=Fy8,
                Fx_eq9=Fx9, Fy_eq9=Fy9)


def process_dataset(files_dict: dict,
                    tethered: list,
                    ref_idx: int,
                    n_zero: int,
                    tau_sh_label: str,
                    exclude_beads: list = None,
                    fail_times: dict = None,
                    out_prefix: str = '') -> pd.DataFrame:
    """
    Process a full tau_sh dataset (all force steps, all beads).

    Identifies the axis parallel to B:
      Using only Eq. 8, the axis // B gives a *lower* apparent force
      (because its effective arm is L_ext + r > L_ext, inflating the variance
       and deflating the Eq. 8 force estimate).
      Once identified, the correct force on that axis uses Eq. 9.

    Returns a tidy DataFrame with one row per (bead, force_set).
    """
    records = []

    for F_set in sorted(files_dict):
        val = files_dict[F_set]
        df, _ = load_data(val)

        for b in tethered:
            t_fail = (fail_times or {}).get(b)
            t_fail_ms = t_fail * 1000.0 if t_fail is not None else None
            res = _analyze_bead(df.copy(), b, ref_idx, n_zero, t_fail_ms)

            # -- Axis identification --------------------------------------
            # Lower Eq. 8 force  ←->  larger variance  ←->  axis // B
            if res['Fx_eq8'] < res['Fy_eq8']:
                F_par  = res['Fx_eq9']   # x is // B  -> correct with Eq. 9
                axis_B = 'x'
            else:
                F_par  = res['Fy_eq9']   # y is // B
                axis_B = 'y'

            records.append({
                'tau_sh':    tau_sh_label,
                'F_set_pN':  F_set,
                'bead':      b,
                'L_ext_um':  res['L_ext_um'],
                'var_x_nm2': res['var_x'],
                'var_y_nm2': res['var_y'],
                'Fx_eq8':    res['Fx_eq8'],
                'Fy_eq8':    res['Fy_eq8'],
                'Fx_eq9':    res['Fx_eq9'],
                'Fy_eq9':    res['Fy_eq9'],
                'F_par_pN':  F_par,
                'axis_B':    axis_B,
            })

    df_out = pd.DataFrame(records)

    # Print summary table (excluding beads flagged for omission)
    _excl = set(exclude_beads) if exclude_beads else set()
    df_calc = df_out[~df_out['bead'].isin(_excl)]

    print(f"\n{'-'*62}")
    print(f"  tau_sh = {tau_sh_label}"
          + (f"  (excluded from stats: beads {sorted(_excl)})" if _excl else ""))
    print(f"{'-'*62}")
    print(f"  {'F_set':>7}  {'F_mean_par':>8}  {'sigma_F':>7}  {'rel%':>7}  n")
    for F_set, grp in df_calc.groupby('F_set_pN'):
        μ   = grp['F_par_pN'].mean()
        σ   = grp['F_par_pN'].std()
        rel = σ / μ * 100 if μ > 0 else 0
        print(f"  {F_set:>7.2f}  {μ:>8.3f}  {σ:>7.3f}  {rel:>6.1f}%  {len(grp)}")

    # Export per-bead variance table to Excel
    _export_variances_excel(df_out, _excl, tau_sh_label, out_prefix)

    return df_calc


def _export_variances_excel(df_out, excl_beads, tau_sh_label, out_prefix):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    beads_sorted  = sorted(df_out['bead'].unique())
    F_sets_sorted = sorted(df_out['F_set_pN'].unique())

    wb = Workbook()
    ws = wb.active
    ws.title = 'Variances'

    # Styles
    font_header  = Font(name='Arial', bold=True, size=10)
    font_normal  = Font(name='Arial', size=10)
    font_excl    = Font(name='Arial', size=10, italic=True, color='888888')
    fill_header  = PatternFill('solid', start_color='2F4F8F', end_color='2F4F8F')
    fill_excl    = PatternFill('solid', start_color='F0F0F0', end_color='F0F0F0')
    fill_alt     = PatternFill('solid', start_color='EEF2FF', end_color='EEF2FF')
    align_center = Alignment(horizontal='center', vertical='center')
    align_right  = Alignment(horizontal='right',  vertical='center')
    thin_side    = Side(style='thin', color='CCCCCC')
    thin_border  = Border(left=thin_side, right=thin_side,
                          top=thin_side, bottom=thin_side)

    # Row 1: title
    ws.merge_cells('A1:B1')
    ws['A1'] = f'Variances per bead [nm²]  —  {tau_sh_label}'
    ws['A1'].font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    ws['A1'].fill = fill_header
    ws['A1'].alignment = align_center
    # extend title across all columns
    n_cols = 2 + len(beads_sorted) * 2
    for c in range(3, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill_header

    # Row 2: bead headers (merged pairs)
    ws['A2'] = 'F_set [pN]'
    ws['A2'].font = font_header
    ws['A2'].fill = PatternFill('solid', start_color='4472C4', end_color='4472C4')
    ws['A2'].font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    ws['A2'].alignment = align_center
    ws['B2'] = 'F_set [pN]'  # placeholder, will be hidden via merge
    ws.merge_cells('A2:B2')

    for i, b in enumerate(beads_sorted):
        col_x = 3 + i * 2
        col_y = col_x + 1
        excl  = b in excl_beads
        label = f'Bead {b}' + (' (excl.)' if excl else '')
        cell  = ws.cell(row=2, column=col_x, value=label)
        cell.font      = Font(name='Arial', bold=True, size=10,
                              color='FFFFFF' if not excl else 'BBBBBB')
        cell.fill      = PatternFill('solid',
                                     start_color='4472C4' if not excl else '888888',
                                     end_color  ='4472C4' if not excl else '888888')
        cell.alignment = align_center
        ws.merge_cells(start_row=2, start_column=col_x,
                       end_row=2,   end_column=col_y)

    # Row 3: x² / y² sub-headers
    ws['A3'] = ''
    for i, b in enumerate(beads_sorted):
        col_x = 3 + i * 2
        for j, lbl in enumerate(['⟨x²⟩', '⟨y²⟩']):
            cell = ws.cell(row=3, column=col_x + j, value=lbl)
            cell.font      = Font(name='Arial', bold=True, size=10)
            cell.alignment = align_center
            cell.border    = thin_border

    ws['A3'].font      = Font(name='Arial', bold=True, size=10)
    ws['A3'].alignment = align_center

    # Data rows
    for r_i, F_set in enumerate(F_sets_sorted):
        row_num = 4 + r_i
        fill_row = fill_alt if r_i % 2 == 0 else PatternFill()

        cell_f = ws.cell(row=row_num, column=1, value=round(F_set, 3))
        cell_f.font      = font_normal
        cell_f.alignment = align_right
        cell_f.fill      = fill_row
        cell_f.border    = thin_border
        ws.merge_cells(start_row=row_num, start_column=1,
                       end_row=row_num,   end_column=2)

        for i, b in enumerate(beads_sorted):
            col_x  = 3 + i * 2
            entry  = df_out[(df_out['F_set_pN'] == F_set) & (df_out['bead'] == b)]
            excl   = b in excl_beads
            f_cell = font_excl if excl else font_normal
            f_bg   = fill_excl if excl else fill_row

            for j, key in enumerate(['var_x_nm2', 'var_y_nm2']):
                val  = round(entry[key].values[0], 1) if len(entry) else None
                cell = ws.cell(row=row_num, column=col_x + j, value=val)
                cell.font      = f_cell
                cell.fill      = f_bg
                cell.alignment = align_right
                cell.border    = thin_border
                if val is not None:
                    cell.number_format = '#,##0.0'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 0.5
    for i in range(len(beads_sorted) * 2):
        ws.column_dimensions[get_column_letter(3 + i)].width = 11

    # Freeze panes below header rows
    ws.freeze_panes = 'A4'

    fname = str(PRACTICAL / f'{out_prefix}_variances.xlsx')
    wb.save(fname)
    print(f"  -> saved {fname}")


# ------------------------------------------------------------------------------
# 5.  PLOTS (Parts 2.1 – 2.3)
# ------------------------------------------------------------------------------

def plot_F_Lext(df: pd.DataFrame, tau_sh_label: str, out_prefix: str = ''):
    """F_par vs L_ext with WLC overlay  (per bead)."""
    fig, ax = plt.subplots(figsize=(7, 5))

    cmap = plt.get_cmap('tab20')
    for i, (b, grp) in enumerate(df.groupby('bead')):
        grp_s = grp.sort_values('F_set_pN')
        ax.plot(grp_s['F_set_pN'], grp_s['L_ext_um'],
                color=cmap(i), marker='o', ms=5, lw=1.2, alpha=0.8,
                label=f'Bead {b}', zorder=3)

    F_wlc = np.logspace(-1.5, 0.9, 300)
    L_wlc = np.array([wlc_L_ext(f) for f in F_wlc])
    ax.plot(F_wlc, L_wlc, 'k--', lw=1.8, label='WLC theory', zorder=4)

    ax.set_xlabel('Expected force  [pN]')
    ax.set_ylabel(r'$L_\mathrm{ext}$  [$\mu$m]')
    ax.set_title(f'Force vs extension — $\\tau_{{sh}}$ = {tau_sh_label}')
    ax.legend(fontsize=8)
    plt.tight_layout()
    out = str(PRACTICAL / f"{out_prefix}_F_Lext_{tau_sh_label.replace(' ','_')}.svg")
    plt.savefig(out, bbox_inches='tight')
    plt.show()
    plt.close()
    print(f"  -> saved {out}")


def plot_var_x_vs_F(df: pd.DataFrame, tau_sh_label: str, out_prefix: str = ''):
    """x-variance vs expected force (per bead)."""
    fig, ax = plt.subplots(figsize=(7, 5))
    cmap = plt.get_cmap('tab20')
    for i, (b, grp) in enumerate(df.groupby('bead')):
        grp_s = grp.sort_values('F_set_pN')
        ax.plot(grp_s['F_set_pN'], grp_s['var_x_nm2'],
                color=cmap(i), marker='o', ms=5, lw=1.2, alpha=0.8,
                label=f'Bead {b}', zorder=3)
    ax.set_xlabel('Expected force  [pN]')
    ax.set_ylabel(r'$\langle x^2 \rangle$  [nm²]')
    ax.set_yscale('log')
    ax.set_title(f'x-variance vs force — $\\tau_{{sh}}$ = {tau_sh_label}')
    ax.legend(fontsize=8)
    plt.tight_layout()
    out = str(PRACTICAL / f"{out_prefix}_var_x_{tau_sh_label.replace(' ','_')}.svg")
    plt.savefig(out, bbox_inches='tight')
    plt.show()
    plt.close()
    print(f"  -> saved {out}")


def plot_var_y_vs_F(df: pd.DataFrame, tau_sh_label: str, out_prefix: str = ''):
    """y-variance vs expected force (per bead)."""
    fig, ax = plt.subplots(figsize=(7, 5))
    cmap = plt.get_cmap('tab20')
    for i, (b, grp) in enumerate(df.groupby('bead')):
        grp_s = grp.sort_values('F_set_pN')
        ax.plot(grp_s['F_set_pN'], grp_s['var_y_nm2'],
                color=cmap(i), marker='o', ms=5, lw=1.2, alpha=0.8,
                label=f'Bead {b}', zorder=3)
    ax.set_xlabel('Expected force  [pN]')
    ax.set_ylabel(r'$\langle y^2 \rangle$  [nm²]')
    ax.set_yscale('log')
    ax.set_title(f'y-variance vs force — $\\tau_{{sh}}$ = {tau_sh_label}')
    ax.legend(fontsize=8)
    plt.tight_layout()
    out = str(PRACTICAL / f"{out_prefix}_var_y_{tau_sh_label.replace(' ','_')}.svg")
    plt.savefig(out, bbox_inches='tight')
    plt.show()
    plt.close()
    print(f"  -> saved {out}")


def plot_Fx_vs_z(df: pd.DataFrame, tau_sh_label: str, out_prefix: str = ''):
    """Measured force from x-variance (Fx_eq9) vs magnet height, with per-z mean."""
    fig, ax = plt.subplots(figsize=(7, 5))
    cmap = plt.get_cmap('tab20')
    for i, (b, grp) in enumerate(df.groupby('bead')):
        grp_s = grp.copy()
        grp_s['z_mm'] = grp_s['F_set_pN'].apply(z_from_F)
        grp_s = grp_s.sort_values('z_mm')
        ax.plot(grp_s['z_mm'], grp_s['Fx_eq9'],
                color=cmap(i), marker='o', ms=4, lw=1.0, alpha=0.7,
                label=f'Bead {b}')
    # Mean across beads per force step
    means = df.groupby('F_set_pN')['Fx_eq9'].mean().reset_index()
    means['z_mm'] = means['F_set_pN'].apply(z_from_F)
    means = means.sort_values('z_mm')
    ax.plot(means['z_mm'], means['Fx_eq9'], 'k-', lw=2.0, label='Mean', zorder=5)
    ax.set_xlabel('Magnet height  $z$  [mm]')
    ax.set_ylabel('$F_x$  [pN]')
    ax.set_title(f'Force from x-variance vs magnet height — $\\tau_{{sh}}$ = {tau_sh_label}')
    ax.legend(fontsize=8)
    plt.tight_layout()
    out = str(PRACTICAL / f"{out_prefix}_Fx_vs_z_{tau_sh_label.replace(' ','_')}.svg")
    plt.savefig(out, bbox_inches='tight')
    plt.show()
    plt.close()
    print(f"  -> saved {out}")


def plot_Fy_vs_z(df: pd.DataFrame, tau_sh_label: str, out_prefix: str = ''):
    """Measured force from y-variance (Fy_eq9) vs magnet height, with per-z mean."""
    fig, ax = plt.subplots(figsize=(7, 5))
    cmap = plt.get_cmap('tab20')
    for i, (b, grp) in enumerate(df.groupby('bead')):
        grp_s = grp.copy()
        grp_s['z_mm'] = grp_s['F_set_pN'].apply(z_from_F)
        grp_s = grp_s.sort_values('z_mm')
        ax.plot(grp_s['z_mm'], grp_s['Fy_eq9'],
                color=cmap(i), marker='o', ms=4, lw=1.0, alpha=0.7,
                label=f'Bead {b}')
    means = df.groupby('F_set_pN')['Fy_eq9'].mean().reset_index()
    means['z_mm'] = means['F_set_pN'].apply(z_from_F)
    means = means.sort_values('z_mm')
    ax.plot(means['z_mm'], means['Fy_eq9'], 'k-', lw=2.0, label='Mean', zorder=5)
    ax.set_xlabel('Magnet height  $z$  [mm]')
    ax.set_ylabel('$F_y$  [pN]')
    ax.set_title(f'Force from y-variance vs magnet height — $\\tau_{{sh}}$ = {tau_sh_label}')
    ax.legend(fontsize=8)
    plt.tight_layout()
    out = str(PRACTICAL / f"{out_prefix}_Fy_vs_z_{tau_sh_label.replace(' ','_')}.svg")
    plt.savefig(out, bbox_inches='tight')
    plt.show()
    plt.close()
    print(f"  -> saved {out}")


def plot_F_ratio_vs_z(df: pd.DataFrame, tau_sh_label: str, out_prefix: str = ''):
    """F_par / F_set vs magnet height, with reference lines, per-axis means and std bands."""
    fig, ax = plt.subplots(figsize=(7, 5))

    df = df.copy()
    df['z_mm']    = df['F_set_pN'].apply(z_from_F)
    df['ratio_x'] = df['Fx_eq9'] / df['F_set_pN']
    df['ratio_y'] = df['Fy_eq9'] / df['F_set_pN']

    # Mean ± std per force step
    stats = (df.groupby('F_set_pN')[['ratio_x', 'ratio_y', 'z_mm']]
               .agg(['mean', 'std'])
               .reset_index())
    stats.columns = ['F_set', 'rx_mean', 'rx_std', 'ry_mean', 'ry_std', 'z_mean', 'z_std']
    stats = stats.sort_values('z_mean')

    z = stats['z_mean'].values

    # x-axis (short)
    ax.errorbar(z, stats['rx_mean'], yerr=stats['rx_std'],
                fmt='b-o', lw=2.0, ms=5, capsize=4, elinewidth=1.5, capthick=1.5,
                label='Mean $F_x/F_{set}$ ± SD  (short axis)', zorder=5)

    # y-axis (long)
    ax.errorbar(z, stats['ry_mean'], yerr=stats['ry_std'],
                fmt='r-o', lw=2.0, ms=5, capsize=4, elinewidth=1.5, capthick=1.5,
                label='Mean $F_y/F_{set}$ ± SD  (long axis)', zorder=5)

    # Reference lines
    ax.axhline(1.0, color='black', lw=1.2, ls='-',  label='$y = 1$  (exact)')
    ax.axhline(1.1, color='black', lw=1.2, ls='--', label='10 % overestimation')

    ax.set_xscale('log')
    ax.set_xlabel('Magnet height  $z$  [mm]  (log scale)')
    ax.set_ylabel('$F_\\mathrm{meas}\ /\ F_\\mathrm{set}$')
    ax.set_title(f'Force ratio vs magnet height — $\\tau_{{sh}}$ = {tau_sh_label}')
    ax.legend(fontsize=8)
    plt.tight_layout()
    out = str(PRACTICAL / f"{out_prefix}_F_ratio_vs_z_{tau_sh_label.replace(' ','_')}.svg")
    plt.savefig(out, bbox_inches='tight')
    plt.show()
    plt.close()
    print(f"  -> saved {out}")


def plot_F_z_combined(dfs: dict, out_prefix: str = ''):
    """F_par vs magnet position z — fit + theory curve."""
    fig, ax = plt.subplots(figsize=(8, 5))

    # Theory curve
    z_arr = np.linspace(0.1, 13, 400)
    ax.plot(z_arr, F_z(z_arr), 'k--', lw=1.5, label='F(z) theory', zorder=5)

    colors = {'1 ms': 'steelblue', '10 ms': 'darkorange', '39 ms': 'seagreen',
              '0.4 ms': 'steelblue'}

    def _double_exp(z, a, b, c, d):
        return a * np.exp(-b * z) + c * np.exp(-d * z)

    for tau_label, df in dfs.items():
        col = colors.get(tau_label, 'gray')

        # Collect (z_mag, mean F_par) per force step for the fit
        z_pts, F_pts = [], []
        first = True
        for F_set, grp in df.groupby('F_set_pN'):
            z_mag  = z_from_F(F_set)
            F_vals = grp['F_par_pN'].values
            lbl    = f'tau_sh = {tau_label}' if first else None
            ax.scatter([z_mag]*len(F_vals), F_vals,
                       color=col, s=22, alpha=0.75, label=lbl, zorder=3)
            z_pts.append(z_mag)
            F_pts.append(np.nanmean(F_vals[np.isfinite(F_vals)]))
            first = False

        # Fit double-exponential to the per-step means (drop NaN/inf)
        z_pts = np.array(z_pts)
        F_pts = np.array(F_pts)
        ok = np.isfinite(F_pts) & np.isfinite(z_pts)
        z_pts, F_pts = z_pts[ok], F_pts[ok]
        try:
            p0 = [22.4, 1.46, 52.3, 0.69]   # start from theory values
            popt, _ = curve_fit(_double_exp, z_pts, F_pts, p0=p0,
                                bounds=(0, np.inf), maxfev=10000)
            z_fit = np.linspace(z_pts.min() * 0.95, z_pts.max() * 1.05, 400)
            ax.plot(z_fit, _double_exp(z_fit, *popt), color='crimson', lw=3,
                    label=(f'fit {tau_label}: '
                           f'${popt[0]:.2f}\\,e^{{-{popt[1]:.3f}z}}'
                           f' + {popt[2]:.2f}\\,e^{{-{popt[3]:.3f}z}}$'), zorder=4)
            print(f"  Fit ({tau_label}): F(z) = {popt[0]:.3f}*exp(-{popt[1]:.4f}*z)"
                  f" + {popt[2]:.3f}*exp(-{popt[3]:.4f}*z)")
        except RuntimeError as e:
            print(f"  Fit failed for {tau_label}: {e}")

    ax.set_yscale('log')
    ax.set_xlabel('Magnet position  $z$  [mm]')
    ax.set_ylabel('Force  [pN]  (log scale)')
    ax.set_title('Force vs magnet position — fit vs theory')
    ax.legend(fontsize=8)
    plt.tight_layout()
    out = str(PRACTICAL / f"{out_prefix}_F_z_combined.svg")
    plt.savefig(out, bbox_inches='tight')
    plt.show()
    plt.close()
    print(f"  -> saved {out}")


def plot_axis_identification(df: pd.DataFrame, tau_sh_label: str, out_prefix: str = ''):
    """
    Show Fx_eq8 and Fy_eq8 vs F_set on a log-log plot.
    The axis // B will systematically underestimate with Eq. 8
    (sits below the F_set diagonal).
    """
    fig, ax = plt.subplots(figsize=(7, 5))
    cmap = plt.get_cmap('tab20')

    for i, (b, grp) in enumerate(df.groupby('bead')):
        c = cmap(i)
        ax.scatter(grp['F_set_pN'], grp['Fx_eq8'],
                   color=c, marker='o', s=30, label=f'B{b} x (Eq.8)')
        ax.scatter(grp['F_set_pN'], grp['Fy_eq8'],
                   color=c, marker='s', s=30, alpha=0.6, label=f'B{b} y (Eq.8)')

    f_rng = [df['F_set_pN'].min()*0.8, df['F_set_pN'].max()*1.2]
    ax.plot(f_rng, f_rng, 'k-', lw=1, label='F_set (ideal)')

    ax.set_xscale('log'); ax.set_yscale('log')
    ax.set_xlabel(r'$F_\mathrm{set}$  [pN]')
    ax.set_ylabel(r'$F_\mathrm{measured}$ (Eq. 8)  [pN]')
    ax.set_title(f'Axis $\\parallel$ B identification — $\\tau_{{sh}}$ = {tau_sh_label}\n'
                 f'($\\parallel$ B axis gives lower Eq. 8 force; correct with Eq. 9)')
    ax.legend(fontsize=7, ncol=2)
    plt.tight_layout()
    out = str(PRACTICAL / f"{out_prefix}_axis_id_{tau_sh_label.replace(' ','_')}.svg")
    plt.savefig(out, bbox_inches='tight')
    plt.show()
    plt.close()
    print(f"  -> saved {out}")


# ------------------------------------------------------------------------------
# 6.  SHUTTER-TIME NORMALIZATION  (Analysis step 4)
# ------------------------------------------------------------------------------

def shutter_normalization(df_1ms: pd.DataFrame,
                          df_10ms: pd.DataFrame,
                          df_39ms: pd.DataFrame):
    """
    Normalize F(10 ms) and F(39 ms) by F(1 ms) per bead per force.
    Plot against  u = t_{c,x} / tau_sh  and overlay theory.

    Theory (from Eq. 11):
        F_meas(tau_sh) / F_true  ~  π / (2·arctan(π·u))
    where  u = t_{c,x} / tau_sh.
    """
    fig, ax = plt.subplots(figsize=(8, 5))
    colors = {'10 ms': 'darkorange', '39 ms': 'seagreen'}
    tau_map = {'10 ms': 10e-3, '39 ms': 39e-3}

    for tau_label, df_tau in [('10 ms', df_10ms), ('39 ms', df_39ms)]:
        tau = tau_map[tau_label]
        col = colors[tau_label]
        first = True

        for b in df_1ms['bead'].unique():
            ref_b = (df_1ms[df_1ms['bead'] == b]
                     .set_index('F_set_pN')['F_par_pN'])
            tau_b = (df_tau[df_tau['bead'] == b]
                     .set_index('F_set_pN')['F_par_pN'])
            Lext_b = (df_1ms[df_1ms['bead'] == b]
                      .set_index('F_set_pN')['L_ext_um'])

            for F_set in ref_b.index:
                if F_set not in tau_b.index:
                    continue
                F_ref  = ref_b[F_set]
                F_tau  = tau_b[F_set]
                L_ext  = Lext_b[F_set]
                tcx    = t_cx(F_set, L_ext)
                u      = tcx / tau
                ratio  = F_tau / F_ref

                lbl = f'tau_sh = {tau_label}' if first else None
                ax.scatter(u, ratio, color=col, s=25, alpha=0.75,
                           label=lbl, zorder=3)
                first = False

    # Theory curve
    u_arr   = np.logspace(-0.5, 2.2, 500)
    theory  = np.pi / (2.0 * np.arctan(np.pi * u_arr))
    ax.plot(u_arr, theory, 'k-', lw=2, label='Theory  π/(2·arctan(π·u))', zorder=4)

    # 10 % overestimation line
    ax.axhline(1.1, color='red', ls='--', lw=1.2, label='10 % overestimation', zorder=2)

    # Find crossing analytically
    u_10 = brentq(lambda u: np.pi / (2*np.arctan(np.pi*u)) - 1.1, 0.1, 200.0)
    ax.axvline(u_10, color='red', ls=':', lw=1.5, zorder=2,
               label=f'Crossing at u = {u_10:.2f}')

    ax.set_xscale('log')
    ax.set_xlabel(r'$t_{c,x} \,/\, \tau_{sh}$')
    ax.set_ylabel(r'$F_\mathrm{meas}(\tau_{sh}) \;/\; F_\mathrm{meas}(1\,\mathrm{ms})$')
    ax.set_title('Shutter-time normalization (Analysis step 4)')
    ax.legend(fontsize=9)
    ax.set_ylim(0.9, None)
    plt.tight_layout()
    plt.savefig(str(PRACTICAL / 'shutter_normalization.svg'), bbox_inches='tight')
    plt.show()
    plt.close()
    print(f"\n  10 % crossing:  t_c,x/tau_sh = {u_10:.2f}")
    print(f"  -> tau_sh must be < t_c,x / {u_10:.2f}  for <10 % overestimation")
    return u_10


# ------------------------------------------------------------------------------
# 7.  STATISTICAL ANALYSIS  (Analysis steps 5 & 6)
# ------------------------------------------------------------------------------

def _force_from_short_file(fpath: str,
                           bead_idx: int,
                           ref_idx: int,
                           L_ext_um: float) -> float:
    """
    Extract F_par from a short measurement file where L_ext is known externally.
    (Used for 2.4 files which have no zero-force period.)
    """
    df, _ = load_data(fpath)

    for ax in ('x', 'y'):
        df[f'd{ax}'] = (df[f'{ax}{bead_idx}'] - df[f'{ax}{ref_idx}']) * 1e3  # nm

    x_c, _ = sigma_clip(df['dx'].values)
    y_c, _ = sigma_clip(df['dy'].values)
    vx = pop_variance(x_c)
    vy = pop_variance(y_c)

    Fx8 = force_eq8(vx, L_ext_um)
    Fy8 = force_eq8(vy, L_ext_um)
    return force_eq9(vx, L_ext_um) if Fx8 < Fy8 else force_eq9(vy, L_ext_um)


def stat_analysis(files_100pts: list,
                  files_50pts: list,
                  tethered: list,
                  ref_idx: int,
                  L_ext_by_bead: dict,
                  df_long_1pN: pd.DataFrame):
    """
    Steps 5 & 6 of the analysis.

    Parameters
    ----------
    files_100pts   : list of 10 file paths, 100-frame acquisitions at 1 pN
    files_50pts    : list of 10 file paths, 50-frame acquisitions at 1 pN
    tethered       : list of bead 1-based indices
    ref_idx        : reference bead index
    L_ext_by_bead  : {bead_idx: L_ext_um}  from the 1 ms dataset at 1 pN
    df_long_1pN    : rows from the 1 ms dataset at F_set = 1 pN (for comparison)
    """

    # -- Step 5 ----------------------------------------------------------------
    print(f"\n{'='*62}")
    print(f"  STEP 5 — Statistical noise (100 and 50 points)")
    print(f"{'='*62}")

    for n_pts, files in [(100, files_100pts), (50, files_50pts)]:
        print(f"\n  --- {n_pts} consecutive points ---")
        print(f"  {'Bead':>5}  {'F_mean  [pN]':>10}  {'sigma_F':>8}  {'rel%':>7}")

        for b in tethered:
            L_ext = L_ext_by_bead[b]
            forces = [_force_from_short_file(f, b, ref_idx, L_ext)
                      for f in files]
            μ   = np.mean(forces)
            σ   = np.std(forces)
            rel = σ / μ * 100 if μ > 0 else 0
            print(f"  {b:>5d}  {μ:>10.3f}  {σ:>8.3f}  {rel:>6.1f}%")

    # -- Step 6  — Concatenate 10 × 100 pts -----------------------------------
    print(f"\n{'='*62}")
    print(f"  STEP 6 — Concatenated 10×100 pts vs long measurement")
    print(f"{'='*62}")
    print(f"  {'Bead':>5}  {'F_concat':>10}  {'F_long':>10}  {'Δ [%]':>8}")

    for b in tethered:
        L_ext = L_ext_by_bead[b]

        # Concatenate all 10 × 100-pt traces for bead b
        x_all, y_all = [], []
        for fpath in files_100pts:
            df, _ = load_data(fpath)
            x_all.extend((df[f'x{b}'] - df[f'x{ref_idx}']).values * 1e3)
            y_all.extend((df[f'y{b}'] - df[f'y{ref_idx}']).values * 1e3)

        x_all = np.array(x_all);  y_all = np.array(y_all)
        vx = pop_variance(sigma_clip(x_all)[0])
        vy = pop_variance(sigma_clip(y_all)[0])
        Fx8 = force_eq8(vx, L_ext); Fy8 = force_eq8(vy, L_ext)
        F_concat = (force_eq9(vx, L_ext) if Fx8 < Fy8
                    else force_eq9(vy, L_ext))

        # Long measurement reference
        row = df_long_1pN[df_long_1pN['bead'] == b]
        F_long = row['F_par_pN'].values[0] if len(row) > 0 else np.nan
        delta  = (F_concat / F_long - 1) * 100 if not np.isnan(F_long) else np.nan
        print(f"  {b:>5d}  {F_concat:>10.3f}  {F_long:>10.3f}  {delta:>7.1f}%")


# ------------------------------------------------------------------------------
# 7b.  OVERVIEW TRACES PLOT
# ------------------------------------------------------------------------------

def plot_traces_overview(df_full: pd.DataFrame,
                         magnet_steps: list,
                         ref_idx: int,
                         beads_to_plot: list,
                         settled_ms: dict = None,
                         out_prefix: str = ''):
    """
    x, y, z position traces over time for a selection of beads.
    Yellow bands mark each force-step window used in the analysis
    (skips step 0 = zero-force and step 1 = stretch).
    """
    t_s = df_full['time_ms'].values / 1e3

    fig, axes = plt.subplots(3, 1, figsize=(14, 8), sharex=True)

    cmap = plt.get_cmap('tab20')
    for i, b in enumerate(beads_to_plot):
        col = cmap(i)
        for ax, coord in zip(axes, ('x', 'y', 'z')):
            drift_corr = (df_full[f'{coord}{b}'] - df_full[f'{coord}{ref_idx}']).values * 1e3
            ax.plot(t_s, drift_corr, lw=0.4, color=col, alpha=0.7, label=f'Bead {b}')

    # Two lines per step: orange = magnet starts moving, green = magnet settled
    # Grey band marks the excluded (still-moving) interval
    t_cursor = 0.0
    for i, (dur_s, z_mm) in enumerate(magnet_steps):
        t_end = t_cursor + dur_s
        t_settled_ms = (settled_ms or {}).get(i)
        t_settled = (t_settled_ms / 1000.0) if t_settled_ms is not None else t_cursor

        for ax in axes:
            ax.axvline(t_cursor,  color='orange',    lw=1.5, ls='--', zorder=5)
            ax.axvline(t_settled, color='limegreen', lw=1.5, ls='--', zorder=5)
            if t_settled > t_cursor + 0.01:
                ax.axvspan(t_cursor, t_settled, color='gray', alpha=0.15, zorder=4)
        t_cursor = t_end

    for ax, ylabel in zip(axes, [r'$\Delta x$  [nm]', r'$\Delta y$  [nm]',
                                  r'$\Delta z$  [nm]']):
        ax.set_ylabel(ylabel)
        ax.axhline(0, color='k', lw=0.4, ls='--')

    axes[2].set_xlabel('Time  [s]')
    axes[0].set_title('Drift-corrected bead positions  '
                      '(orange = magnet starts moving, green = settled; grey = excluded)')
    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc='upper right', fontsize=7, ncol=2)
    plt.tight_layout()
    bead_label = '_'.join(f'b{b}' for b in beads_to_plot)
    out = str(PRACTICAL / f'{out_prefix}_traces_overview_{bead_label}.svg')
    plt.savefig(out, bbox_inches='tight')
    plt.close()
    print(f"  -> saved {out}")


# ------------------------------------------------------------------------------
# 8.  THEORETICAL PLOTS  (Preparatory questions 1–4 from the PDF)
# ------------------------------------------------------------------------------

# Force values used in preparatory questions (pN)
_PREP_FORCES = np.array([0.1, 0.2, 0.3, 0.4, 0.8, 1.0, 3.0, 5.0, 6.0])


def theo_plot_F_vs_Lext():
    """
    Preparatory Q1 — F(L_ext): WLC force vs extension (Eq. 12).
    x-axis: L_ext [µm]  (log scale)
    y-axis: Force [pN]
    """
    # Range: from L_ext at 0.01 pN up to L_ext at 100 pN (avoid divergence)
    F_max_plot = 100.0
    L_max_um   = wlc_L_ext(F_max_plot)
    L_arr = np.linspace(wlc_L_ext(0.01), L_max_um, 800)  # µm
    F_arr = np.array([wlc_force(l * 1e-6) for l in L_arr])

    fig, ax = plt.subplots(figsize=(7, 5))
    ax.plot(L_arr, F_arr, 'b-', lw=2, label='WLC  (Eq. 12)')

    # Mark the preparatory-question force values
    for F_pN in _PREP_FORCES:
        L_um = wlc_L_ext(F_pN)
        ax.scatter(L_um, F_pN, color='red', s=40, zorder=5)
        ax.annotate(f'{F_pN} pN', (L_um, F_pN),
                    textcoords='offset points', xytext=(5, 3), fontsize=7)

    L_min_um = wlc_L_ext(0.05)   # just below the lowest force point
    ax.set_xscale('log')
    ax.set_xlim(L_min_um * 0.9, L_max_um * 1.01)
    ax.set_ylim(0, 8)
    ax.set_xlabel(r'$L_\mathrm{ext}$  [$\mu$m]')
    ax.set_ylabel('Force  [pN]')
    ax.set_title('WLC: Force vs Extension  (Eq. 12)')
    ax.legend(fontsize=9)
    plt.tight_layout()
    out = str(THEORETICAL / 'theo_F_vs_Lext.svg')
    plt.savefig(out, bbox_inches='tight')
    plt.close()
    print(f'  -> saved {out}')


def theo_plot_tcx_vs_F():
    """
    Preparatory Q2 — t_{c,x}(F): characteristic time vs force (Eq. 7).
    x-axis: Force [pN]  (log scale)
    y-axis: t_{c,x} [s]
    Also marks 1/f_ac = 1/58 Hz reference line.
    """
    F_arr   = np.logspace(-1.5, 0.9, 400)  # pN
    L_arr   = np.array([wlc_L_ext(f) for f in F_arr])
    tcx_arr = np.array([t_cx(f, l) for f, l in zip(F_arr, L_arr)])

    fig, ax = plt.subplots(figsize=(7, 5))
    ax.plot(F_arr, tcx_arr, 'b-', lw=2, label=r'$t_{c,x} = \gamma L_\mathrm{ext}/F$  (Eq. 7)')
    ax.axhline(1.0 / 58.0, color='red', ls='--', lw=1.5,
               label=r'$1/f_{ac}$ = 1/58 Hz  ≈ 17 ms')

    # Mark the preparatory-question force values
    for F_pN in _PREP_FORCES:
        L_um = wlc_L_ext(F_pN)
        tc   = t_cx(F_pN, L_um)
        ax.scatter(F_pN, tc, color='darkorange', s=40, zorder=5)

    ax.set_yscale('log')
    ax.set_xlabel('Force  [pN]')
    ax.set_ylabel(r'$t_{c,x}$  [s]  (log scale)')
    ax.set_title(r'Characteristic time $t_{c,x}$ vs Force  (Eq. 7)')
    ax.legend(fontsize=9)
    plt.tight_layout()
    out = str(THEORETICAL / 'theo_tcx_vs_F.svg')
    plt.savefig(out, bbox_inches='tight')
    plt.close()
    print(f'  -> saved {out}')


def theo_plot_F_vs_magnet():
    """
    Preparatory Q4 — F(z): force vs magnet position (double-exponential calibration).
    x-axis: magnet position z [mm]
    y-axis: Force [pN]
    Marks the positions corresponding to the nine preparatory forces.
    """
    z_arr = np.linspace(0.1, 13.0, 400)
    F_arr = F_z(z_arr)

    fig, ax = plt.subplots(figsize=(7, 5))
    ax.plot(z_arr, F_arr, 'b-', lw=2,
            label=r'$F_z = 22.38\,e^{-1.458z} + 52.30\,e^{-0.691z}$')

    for F_pN in _PREP_FORCES:
        z_mm = z_from_F(F_pN)
        ax.scatter(z_mm, F_pN, color='red', s=40, zorder=5)
        ax.annotate(f'{F_pN} pN', (z_mm, F_pN),
                    textcoords='offset points', xytext=(4, 2), fontsize=7)

    ax.set_xlabel('Magnet position  $z$  [mm]')
    ax.set_ylabel('Force  [pN]')
    ax.set_title('Force vs Magnet Position  (Preparatory Q4)')
    ax.legend(fontsize=9)
    plt.tight_layout()
    out = str(THEORETICAL / 'theo_F_vs_magnet.svg')
    plt.savefig(out, bbox_inches='tight')
    plt.close()
    print(f'  -> saved {out}')


def theo_plot_varx_vs_F():
    """
    Preparatory Q3 — <x²>(F): theoretical x-variance vs force (Eq. 8).
    x-axis: Force [pN]  (log scale)
    y-axis: <x²> [nm²]
    """
    F_arr   = np.logspace(-1.5, 0.9, 400)
    L_arr   = np.array([wlc_L_ext(f) for f in F_arr])
    vx_arr  = np.array([var_x_theory(f, l) for f, l in zip(F_arr, L_arr)])

    fig, ax = plt.subplots(figsize=(7, 5))
    ax.plot(F_arr, vx_arr, 'b-', lw=2,
            label=r'$\langle x^2\rangle = k_BT\,L_\mathrm{ext}/F$  (Eq. 8)')

    for F_pN in _PREP_FORCES:
        L_um = wlc_L_ext(F_pN)
        ax.scatter(F_pN, var_x_theory(F_pN, L_um), color='red', s=40, zorder=5)

    ax.set_xscale('log')
    ax.set_yscale('log')
    ax.set_xlabel('Force  [pN]  (log scale)')
    ax.set_ylabel(r'$\langle x^2 \rangle$  [nm²]  (log scale)')
    ax.set_title(r'Theoretical $\langle x^2\rangle$ vs Force  (Eq. 8)')
    ax.legend(fontsize=9)
    plt.tight_layout()
    out = str(THEORETICAL / 'theo_varx_vs_F.svg')
    plt.savefig(out, bbox_inches='tight')
    plt.close()
    print(f'  -> saved {out}')


def theo_plot_vary_vs_F():
    """
    Preparatory Q3 — <y²>(F): theoretical y-variance vs force (Eq. 9).
    x-axis: Force [pN]  (log scale)
    y-axis: <y²> [nm²]
    """
    F_arr   = np.logspace(-1.5, 0.9, 400)
    L_arr   = np.array([wlc_L_ext(f) for f in F_arr])
    vy_arr  = np.array([var_y_theory(f, l) for f, l in zip(F_arr, L_arr)])

    fig, ax = plt.subplots(figsize=(7, 5))
    ax.plot(F_arr, vy_arr, 'r-', lw=2,
            label=r'$\langle y^2\rangle = k_BT\,(L_\mathrm{ext}+R)/F$  (Eq. 9)')

    for F_pN in _PREP_FORCES:
        L_um = wlc_L_ext(F_pN)
        ax.scatter(F_pN, var_y_theory(F_pN, L_um), color='darkred', s=40, zorder=5)

    ax.set_xscale('log')
    ax.set_yscale('log')
    ax.set_xlabel('Force  [pN]  (log scale)')
    ax.set_ylabel(r'$\langle y^2 \rangle$  [nm²]  (log scale)')
    ax.set_title(r'Theoretical $\langle y^2\rangle$ vs Force  (Eq. 9)')
    ax.legend(fontsize=9)
    plt.tight_layout()
    out = str(THEORETICAL / 'theo_vary_vs_F.svg')
    plt.savefig(out, bbox_inches='tight')
    plt.close()
    print(f'  -> saved {out}')


def theo_table_F_vs_magnet():
    """
    Preparatory Q4 — table of Force (pN) vs magnet position (mm).
    Saves as CSV and prints to console.
    """
    rows = []
    for F_pN in _PREP_FORCES:
        z_mm   = z_from_F(F_pN)
        L_um   = wlc_L_ext(F_pN)
        tc_s   = t_cx(F_pN, L_um)
        vx_nm2 = var_x_theory(F_pN, L_um)
        vy_nm2 = var_y_theory(F_pN, L_um)
        rows.append(dict(
            F_pN=F_pN,
            z_mm=round(z_mm, 4),
            L_ext_um=round(L_um, 3),
            t_cx_s=round(tc_s, 4),
            var_x_nm2=round(vx_nm2, 1),
            var_y_nm2=round(vy_nm2, 1),
        ))

    df = pd.DataFrame(rows)
    df.columns = ['Force (pN)', 'Magnet pos z (mm)', 'L_ext (µm)',
                  't_c,x (s)', '<x²> (nm²)', '<y²> (nm²)']

    print(f"\n{'='*72}")
    print("  TABLE: Theoretical values vs Force  (Preparatory questions)")
    print(f"{'='*72}")
    print(df.to_string(index=False))
    print(f"{'='*72}\n")

    out_csv = str(THEORETICAL / 'theo_table_F_vs_magnet.csv')
    df.to_csv(out_csv, index=False)
    print(f'  -> saved {out_csv}')
    return df


def make_theoretical_plots():
    """Generate and save all six theoretical plots to the 'theoretical/' folder."""
    print(f"\n{'='*62}")
    print("  THEORETICAL PLOTS  (Preparatory questions)")
    print(f"{'='*62}")
    theo_plot_F_vs_Lext()
    theo_plot_tcx_vs_F()
    theo_plot_F_vs_magnet()
    theo_plot_varx_vs_F()
    theo_plot_vary_vs_F()
    theo_table_F_vs_magnet()


# ------------------------------------------------------------------------------
# 9.  MAIN — fill in your file paths and run
# ------------------------------------------------------------------------------

def run_experiment(base: Path, n_total: int, tau_sh: str,
                   exclude_beads: list, trace_beads: list,
                   out_prefix: str, fail_times: dict = None):
    """Load, segment, and plot one experiment folder."""
    ref_idx  = n_total
    tethered = list(range(1, n_total - 1))

    print(f"\n{'='*62}")
    print(f"  {out_prefix}  |  tau_sh = {tau_sh}  |  {n_total} beads")
    print(f"  Saving plots to: {PRACTICAL.resolve()}")
    print(f"{'='*62}")

    # Load traces
    print("Loading traces.txt ...")
    df_full, _ = load_data(str(base / 'traces.txt'))

    # Parse magnet script
    magnet_steps = []
    with open(str(base / 'magnet-script.txt')) as fh:
        for line in fh:
            parts = line.strip().split()
            if len(parts) >= 2:
                try:
                    magnet_steps.append((float(parts[0]), float(parts[1])))
                except ValueError:
                    pass

    # Parse magnet history -> settled times
    mh = pd.read_csv(str(base / 'magnet-history.txt'), sep='\t', comment='#',
                     header=None,
                     names=['t', 'target_pos', 'target_speed',
                            'target_rot', 'rot_speed', 'actual_pos', 'actual_rot'])
    SETTLE_TOL = 0.01
    transition_rows = mh[mh['target_pos'] != mh['target_pos'].shift()].index.tolist()
    settled_ms = {}
    for step_i, row_idx in enumerate(transition_rows):
        target = mh.loc[row_idx, 'target_pos']
        after  = mh.loc[row_idx:]
        ok     = after[abs(after['actual_pos'] - target) < SETTLE_TOL]
        settled_ms[step_i] = ok.iloc[0]['t'] * 1000.0 if len(ok) else None

    # Slice into per-force segments
    t_ms = 0.0
    zero_df  = None
    n_zero   = 0
    seg_dict = {}

    for i, (dur_s, z_mm) in enumerate(magnet_steps):
        t_end_ms = t_ms + dur_s * 1000.0
        if i == 0:
            mask    = (df_full['time_ms'] >= t_ms) & (df_full['time_ms'] < t_end_ms)
            zero_df = df_full[mask].copy().reset_index(drop=True)
            n_zero  = len(zero_df)
            print(f"  Zero-force: {dur_s:.0f} s -> {n_zero} frames  (z = {z_mm} mm)")
        elif i == 1:
            print(f"  Stretch step skipped (z = {z_mm} mm)")
        else:
            t_data_start = settled_ms.get(i, t_ms) or t_ms
            mask  = (df_full['time_ms'] >= t_data_start) & (df_full['time_ms'] < t_end_ms)
            chunk = df_full[mask].copy().reset_index(drop=True)
            F_pN  = round(F_z(z_mm), 3)
            seg_dict[F_pN] = pd.concat([zero_df, chunk], ignore_index=True)
            print(f"  Force step {i-1:2d}: z = {z_mm:.2f} mm  ->  F ~ {F_pN:.3f} pN  "
                  f"({len(chunk)} frames, skipped {(t_data_start-t_ms)/1000:.1f} s)")
        t_ms = t_end_ms

    print(f"\n  n_zero = {n_zero}  |  force steps: {len(seg_dict)}")

    # Overview traces
    for b in trace_beads:
        plot_traces_overview(df_full, magnet_steps, ref_idx, [b], settled_ms,
                             out_prefix=out_prefix)

    # Force calibration
    print(f"\n=== Force calibration ({out_prefix}, tau_sh = {tau_sh}) ===")
    df_cal = process_dataset(seg_dict, tethered, ref_idx, n_zero, tau_sh,
                             exclude_beads=exclude_beads, fail_times=fail_times,
                             out_prefix=out_prefix)
    plot_F_Lext(df_cal, tau_sh, out_prefix=out_prefix)
    plot_var_x_vs_F(df_cal, tau_sh, out_prefix=out_prefix)
    plot_var_y_vs_F(df_cal, tau_sh, out_prefix=out_prefix)
    plot_Fx_vs_z(df_cal, tau_sh, out_prefix=out_prefix)
    plot_Fy_vs_z(df_cal, tau_sh, out_prefix=out_prefix)
    plot_F_ratio_vs_z(df_cal, tau_sh, out_prefix=out_prefix)
    plot_F_z_combined({tau_sh: df_cal}, out_prefix=out_prefix)

    return df_cal


def _parse_magnet_script(base: Path):
    """Return list of (duration_s, z_mm) from magnet-script.txt."""
    steps = []
    with open(str(base / 'magnet-script.txt')) as fh:
        for line in fh:
            parts = line.strip().split()
            if len(parts) >= 2:
                try:
                    steps.append((float(parts[0]), float(parts[1])))
                except ValueError:
                    pass
    return steps


def interactive_run():
    """Fully interactive experiment runner."""
    import sys

    # ------------------------------------------------------------------
    # Step 0: theoretical plots (always generated first)
    # ------------------------------------------------------------------
    make_theoretical_plots()

    # ------------------------------------------------------------------
    # Step 1: experiment folder
    # ------------------------------------------------------------------
    exp_name = input("\nWhat is your experiment map called: ").strip()
    base = Path(exp_name)
    if not base.exists():
        print(f"ERROR: folder '{base}' not found.")
        sys.exit(1)

    # ------------------------------------------------------------------
    # Step 2: auto-detect number of beads from traces.txt
    # ------------------------------------------------------------------
    traces_path = base / 'traces.txt'
    df_full, n_total = load_data(str(traces_path))
    ref_idx  = n_total          # last bead is reference (N2 used for drift)
    tethered = list(range(1, n_total - 1))
    print(f"  Detected {n_total} beads total → "
          f"tethered: {tethered},  reference: {n_total - 1} & {n_total}")

    # ------------------------------------------------------------------
    # Step 3: auto-detect shutter time from frame interval in traces.txt
    # ------------------------------------------------------------------
    dt_ms   = float(np.median(np.diff(df_full['time_ms'].values)))
    f_ac_hz = 1000.0 / dt_ms                        # acquisition frequency [Hz]
    tau_default_ms = round(dt_ms, 4)                 # tau_sh = 1/f_ac (zero dead time)
    tau_default_label = (f"{int(tau_default_ms)} ms"
                         if tau_default_ms == int(tau_default_ms)
                         else f"{tau_default_ms} ms")
    print(f"  Detected frame interval: {dt_ms:.3f} ms  →  "
          f"f_ac ≈ {f_ac_hz:.1f} Hz  →  tau_sh = {tau_default_label}")
    override = input(f"  Shutter time tau_sh [{tau_default_label}] "
                     f"(press Enter to accept, or type a value): ").strip()
    tau_sh = override if override else tau_default_label

    # ------------------------------------------------------------------
    # Step 4: set output folder = experiment name, show traces for ALL beads
    # ------------------------------------------------------------------
    global PRACTICAL
    PRACTICAL = Path(__file__).parent.parent / "plots" / "practical" / base.name
    PRACTICAL.mkdir(parents=True, exist_ok=True)
    out_prefix = base.name          # just the last folder segment, e.g. '20260604_1602_exp1'

    magnet_steps = _parse_magnet_script(base)

    # Parse magnet history to get settled times (needed for 2-line markers)
    mh = pd.read_csv(str(base / 'magnet-history.txt'), sep='\t', comment='#',
                     header=None,
                     names=['t', 'target_pos', 'target_speed',
                            'target_rot', 'rot_speed', 'actual_pos', 'actual_rot'])
    SETTLE_TOL = 0.01
    transition_rows = mh[mh['target_pos'] != mh['target_pos'].shift()].index.tolist()
    settled_ms = {}
    for step_i, row_idx in enumerate(transition_rows):
        target = mh.loc[row_idx, 'target_pos']
        after  = mh.loc[row_idx:]
        ok     = after[abs(after['actual_pos'] - target) < SETTLE_TOL]
        settled_ms[step_i] = ok.iloc[0]['t'] * 1000.0 if len(ok) else None

    print(f"\nShowing traces for all {len(tethered)} tethered bead(s) — close each "
          f"plot window to continue...")
    for b in tethered:
        plot_traces_overview(df_full, magnet_steps, ref_idx, [b], settled_ms,
                             out_prefix=out_prefix)

    # ------------------------------------------------------------------
    # Step 5: excluded beads
    # ------------------------------------------------------------------
    excl_raw = input(
        "\nWhich beads should be excluded? "
        "(comma-separated numbers, e.g. '3,7' — press Enter for none): "
    ).strip()
    exclude_beads = (
        [int(x.strip()) for x in excl_raw.split(',') if x.strip()]
        if excl_raw else []
    )
    if exclude_beads:
        print(f"  Excluding beads: {exclude_beads}")

    # ------------------------------------------------------------------
    # Step 6: fail times per bead
    # ------------------------------------------------------------------
    active_beads = [b for b in tethered if b not in exclude_beads]
    print("\nEnter the fail time in ms for each bead "
          "(the time in the experiment at which the DNA tether breaks).")
    print("Enter 0 or leave blank if the bead has no fail time.")
    fail_times = {}
    for b in active_beads:
        val = input(f"  Bead {b} fail time [ms] (0 = none): ").strip()
        try:
            t = float(val)
            if t > 0:
                fail_times[b] = t
        except ValueError:
            pass   # blank or non-numeric → no fail time

    if fail_times:
        print(f"  Fail times set: {fail_times}")

    # ------------------------------------------------------------------
    # Step 7: run analysis — traces already shown, skip inside run_experiment
    # ------------------------------------------------------------------
    run_experiment(
        base          = base,
        n_total       = n_total,
        tau_sh        = tau_sh,
        exclude_beads = exclude_beads,
        trace_beads   = [],     # already shown above
        out_prefix    = out_prefix,
        fail_times    = fail_times,
    )

    print("\nAll done.")


if __name__ == '__main__':
    interactive_run()